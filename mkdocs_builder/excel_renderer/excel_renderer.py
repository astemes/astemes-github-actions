from pathlib import Path
import openpyxl

def render_excel_as_markdown(directory):
    renderer = excel_renderer()
    excel_files = find_excel_files(directory)
    for excel_file in excel_files:
        md_content = renderer.render_as_markdown(excel_file)
        create_md_file_from_xlsx(excel_file, md_content)

def find_excel_files(directory):
    xlsx_files = list(Path(directory).rglob("*.[xX][lL][sS]*"))
    return xlsx_files

def create_md_file_from_xlsx(excel_file, md_content):
    assert isinstance(excel_file,Path)
    md_file = excel_file.with_name(excel_file.stem+".md")
    f = md_file.open("w", encoding = "UTF-8")
    f.write(md_content)
    f.close()


class excel_renderer:
    def render_as_markdown(self, file, sheet_index = 0):
        file_path_str = str(file)
        workbook = openpyxl.load_workbook(filename = file_path_str)
        sheet = workbook.active
        header = self._render_page_header(sheet) +"\n\n"
        content = []
        content.append(self._render_table_headers(sheet))
        content.append(self._render_content(sheet))
        content = "\n".join(content)
        if content.isspace():
            return header
        else:
            return header + content

    def _render_page_header(self,sheet):
        return "# "+sheet.title

    def _render_table_headers(self,sheet):
        first_row = sheet.iter_rows(min_row=1,max_row=1)
        header_items = []
        for row in first_row:
            for cell in row:
                if cell.value:
                    header_items.append(cell.value)
        header_vertical_separator = ["---" for _ in header_items]
        header_items = self._interleave(header_items)
        header_vertical_separator = self._interleave(header_vertical_separator)
        header_items = "".join(header_items)
        header_vertical_separator = "".join(header_vertical_separator)
        header = [header_items,header_vertical_separator]
        return "\n".join(header)

    def _render_content(self,sheet):
        rows = sheet.iter_rows(min_row=2)
        row_items = []
        for row in rows:
            cell_values = []
            for cell in row:
                if cell.value:
                    cell_values.append(str(cell.value))
            cell_values = self._interleave(cell_values)
            row_items.append("".join(cell_values))
        if not row_items:
            return ""
        return "\n".join(row_items)


    def _interleave(self, list, separator = "|"):
        separators = []
        for i in range(len(list)):
            separators.append(separator)
        interleaved_list = [val for pair in zip(separators, list) for val in pair]
        if interleaved_list:
            interleaved_list.append(separator)
        return interleaved_list
